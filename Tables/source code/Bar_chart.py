#!E:\Python Learing\Automation Script\env\Scripts\python.exe
#Import Package
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference

#Load Excel File
wb = load_workbook('pivot_tables.xlsx')
sheet = wb['Report']

min_co=wb.active.min_column
max_co=wb.active.max_column
min_row=wb.active.min_row
max_row=wb.active.max_row

#Define Bar chart
barchart=BarChart()

data=Reference(sheet,
          min_col=min_co+1,
          max_col=max_co,
          min_row=min_row,
          max_row=max_row)

categories=Reference(sheet,
          min_col=min_co,
          max_col=min_co,
          min_row=min_row+1,
          max_row=max_row)

barchart.add_data(data, titles_from_data=True)
barchart.set_categories(categories)
sheet.add_chart(barchart,"B12")

barchart.title="Sales by Product line"
barchart.style = 5
wb.save('barchart.xlsx')