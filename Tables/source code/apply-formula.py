#!E:\Python Learing\Automation Script\env\Scripts\python.exe
#Import Package
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

#Load Excel File
wb = load_workbook('pivot_tables.xlsx')
sheet = wb['Report']

#Get all min max Row and Column
min_co=wb.active.min_column
max_co=wb.active.max_column
min_row=wb.active.min_row
max_row=wb.active.max_row

#Using For loop to auto calculate SUM
for i in range(min_co+1,max_co+1):
    letter=get_column_letter(i)
    sheet[f'{letter}{max_row+1}'] = f'=SUM({letter}{min_row+1}:{letter}{max_row})'
    sheet[f'{letter}{max_row+1}'].style = 'Currency'

wb.save('pivot_Sum_report.xlsx')