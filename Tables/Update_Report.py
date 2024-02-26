#! E:\Python Learing\Automation Script\env\Scripts\python.exe
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
import os
import sys

application_path = os.path.dirname(sys.executable)

#Define Variable
month = input('Introduce month: ')

#Load Excel File
input_path = os.path.join(application_path,'pivot_tables.xlsx')
wb = load_workbook(input_path)
sheet = wb['Report']

#Get all min max Row and Column
min_co=wb.active.min_column
max_co=wb.active.max_column
min_row=wb.active.min_row
max_row=wb.active.max_row

#Define Bar chart
barchart=BarChart()

data=Reference(sheet,
          min_col=min_co+1,
          max_col=max_co,
          min_row=min_row,
          max_row=max_row)

categories=Reference(sheet,
          min_col=min_co,
          max_col=min_co,
          min_row=min_row+1,
          max_row=max_row)

barchart.add_data(data, titles_from_data=True)
barchart.set_categories(categories)
sheet.add_chart(barchart,"B12")

barchart.title="Sales by Product line"
barchart.style = 5

#Using For loop to auto calculate SUM
for i in range(min_co+1,max_co+1):
    letter=get_column_letter(i)
    sheet[f'{letter}{max_row+1}'] = f'=SUM({letter}{min_row+1}:{letter}{max_row})'
    sheet[f'{letter}{max_row+1}'].style = 'Currency'

#Format Cells
sheet['A1'] = 'Sales Report'
sheet['A2'] = month
sheet['A1'].font =Font('Arial',bold=True, size=20)
sheet['A2'].font =Font('Arial',bold=True, size=10)


#Save the Result
output_path = os.path.join(application_path, f'Report_{month}.xlsx')
wb.save(output_path)
print("Report has been updated!!!!")
end = input("Press enter to exit")
